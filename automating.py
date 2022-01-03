import csv
import shutil
import pandas as pd
from openpyxl import load_workbook

                #list of old emails '...@helpinghands.cm'
csv_path = 'helpinghands.csv'
xcel_path = "helpinghands.xlsx"
                #list of new emails '...@handsinhands.org'
new_excel_path = "handsinhands.xlsx"
new_csv_path = "handsinhands.csv"
                #opening csv file in write only
file = open(csv_path, 'w')
                #using csv writer function to prepare query into old csv file
writer = csv.writer(file)

                #data prepared for query
header = ("First Name", "Last Name", "Email", "Phone")
data = [                                #email data is left blank. it will be edited in the generated excel sheet
    ("John", "Doe", "", 458213694),
    ("Mad", "Horse", "", 12033687),
    ("Crazed", "Buffalo", "", 20038792),
    ("Sleeping", "Donkey", "", 200879541),
    ("Drunk", "Stallion", "", 206987632),
    ("Ugly", "Buffoon", "", 487963254),
    ("Laughing", "Dog", "", 412036983),
    ("Swift", "Tortoise", "", 987563200),
    ("Blind", "Eagle", "", 410689752),
    ("Raging", "Tilapia", "", 223698789),
    ("Smart", "Buffoon", "", 236998765),
    ("Handsome", "Gorilla", "", 120039874),
    ("Sluggish", "Antelope", "", 589645632),
    ("Foolish", "Banana", "", 369741013),
    ("Grand", "Kumba", "", 666666666),
    ("Edoudua", "Nonglace", "", 32085000),
    ("Nonchalant", "Pig", "", 700000123),
    ("Blazing", "Rat", "", 203692323),
    ("Flying", "Crocodile", "", 121212121),
    ("Deadly", "Monsoon", "", 322336987),
    ("Swift", "Tiger", "", 989897985),
    ("Dirty", "Girl", "", 785213698),
    ("Ambushed", "Man", "", 218787856),
    ("Scape", "Goat", "", 215696565),
    ("Ngong", "Dog", "", 219874652),
    ("Stale", "Milk", "", 120236987),
    ("Beef", "Stew", "", 123689787),
    ("Monkey", "Kola", "", 458796520),
    ("Drunk", "Driver", "", 123098774),
    ("Okada", "Man", "", 100002365)
]
                #writing prepared data into csv file
writer.writerow(header)
writer.writerows(data)

file.close()
                #generating excel sheet
toExcel = pd.read_csv(csv_path)
toExcel.to_excel(xcel_path, index=False)
                #loading and preparing generated excel sheet for query
workbook = load_workbook(xcel_path)
sheet = workbook.active
                #editing the email column of each data entry
r = 2
while r < sheet.max_row + 1:
    sheet.cell(r, 3).value = sheet.cell(r, 1).value + sheet.cell(r, 2).value + "@helpinghands.cm"
    r += 1

workbook.save(xcel_path)
workbook.close()
                #updating the related csv file
read_file = pd.read_excel(xcel_path, usecols=header)
read_file.to_csv(csv_path, index=None, header=True)
                #copying old csv data into new csv file "@handsinhands.org"
shutil.copy(csv_path, new_csv_path)
                #generating new excel sheet
newExcel = pd.read_csv(new_csv_path)
toExcel.to_excel(new_excel_path, index=False)
                #preparing new excel sheet for query
workbook = load_workbook(new_excel_path)
new_sheet = workbook.active
                #replacing email
o = 2
while o <= new_sheet.max_row:
    new_sheet.cell(o, 3).value = new_sheet.cell(o, 1).value + new_sheet.cell(o, 2).value + "@handsinhands.org"
    o += 1

workbook.save(new_excel_path)

read_new_file = pd.read_excel(new_excel_path, usecols=header)
read_new_file.to_csv(new_csv_path, index=None, header=True)
